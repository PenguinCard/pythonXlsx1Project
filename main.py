from openpyxl import load_workbook, Workbook

import os


def main():
    files = os.listdir()
    xlsx_files = list(filter(lambda name: True if name.find('.xlsx') >= 0 else False, files))
    xlsx_files.sort()

    print(xlsx_files)

    cols = ['A', 'B']

    file_name = "union.xlsx"
    write_wb = Workbook()
    write_ws = write_wb.active

    for xlsx_file in xlsx_files:
        idx = 2
        print(idx, xlsx_file)
        wb = load_workbook(xlsx_file)
        sheet = wb.active

        while any(list(map(lambda col: sheet[col + str(idx)].value is not None, cols))):
            print(list(map(lambda col: str(sheet[col + str(idx)].value), cols)))
            write_ws.append(list(map(lambda col: str(sheet[col + str(idx)].value), cols)))
            idx = idx + 1

    write_wb.save(file_name)


if __name__ == "__main__":
    main()
